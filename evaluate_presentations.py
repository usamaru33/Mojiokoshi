"""
evaluate_presentations.py
=========================
プレゼン評価パイプライン
1. 動画 (Reference/presentation_movies/ 以下) から Whisper で文字起こし
   - 対応形式: .mp4 / .mov / .mkv / .avi およびその大文字拡張子
   - 分割動画の自動統合: 8-1, 8-2, 8-3 → 8_transcript.txt に結合
2. Evaluation Sheet.xlsx の発表者リストと照合
3. 文字起こしテキストを評価プロンプトに渡し、CSVを生成

【使い方】
  python evaluate_presentations.py                     # 文字起こし + 評価プロンプト生成（一括）
  python evaluate_presentations.py --transcribe-only   # 文字起こしのみ
  python evaluate_presentations.py --evaluate-only     # 評価プロンプト生成のみ
  python evaluate_presentations.py --list              # 発表者リストを表示
  python evaluate_presentations.py --model medium      # Whisper モデルを指定
"""

import os
import sys
import csv
import re
import argparse
from pathlib import Path

# ─────────────────────────────────────────
#  パス設定
# ─────────────────────────────────────────
BASE_DIR    = Path(__file__).parent
EVAL_SHEET  = BASE_DIR / "Evaluation Sheet.xlsx"
PRES_MOVIES = BASE_DIR / "Reference" / "presentation_movies"       # 発表動画置き場
PRES_TRANS  = BASE_DIR / "Reference" / "presentation_transcripts"  # 文字起こし置き場
OUTPUT_CSV  = BASE_DIR / "evaluation_results.csv"
WHISPER_MODEL = "base"  # tiny / base / small / medium / large

# 対応する動画拡張子
VIDEO_EXTENSIONS = [
    ".mp4", ".MP4",
    ".mov", ".MOV",
    ".mkv", ".MKV",
    ".avi", ".AVI",
    ".m4v", ".M4V",
    ".webm", ".WEBM",
]

# 無音誤検出とみなす単語
NOISE_WORDS = {"you", "uh", "um", "hmm", "ah", "oh", "okay", "uhh", "uhm"}
# 同一ノイズ単語が全体の何割以上を占めたらゴミとみなす
GARBAGE_RATIO = 0.75
# 末尾ノイズをトリミングする際のウィンドウサイズ（単語数）
TRIM_WINDOW = 30

# ─────────────────────────────────────────
#  ディレクトリ初期化
# ─────────────────────────────────────────
PRES_MOVIES.mkdir(parents=True, exist_ok=True)
PRES_TRANS.mkdir(parents=True, exist_ok=True)


# ─────────────────────────────────────────
#  Evaluation Sheet から発表者リストを読み込む
# ─────────────────────────────────────────
def load_presenter_list(xlsx_path: Path) -> list:
    try:
        import openpyxl
    except ImportError:
        print("[ERROR] openpyxl が未インストールです: pip install openpyxl")
        sys.exit(1)

    print(f"[INFO] Evaluation Sheet を読み込み中: {xlsx_path.name} ...")
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb.active

    # ヘッダー行を探す
    header_row = None
    for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if row[2] and "Presenter" in str(row[2]):
            header_row = i
            break
    if header_row is None:
        print("[ERROR] ヘッダー行が見つかりませんでした")
        sys.exit(1)

    presenters = []
    for row in ws.iter_rows(min_row=header_row + 2, values_only=True):
        week, number, name = row[0], row[1], row[2]
        if not name or name == "Example man":
            continue
        presenters.append({
            "week":         week,
            "number":       number,
            "name":         str(name).strip(),
            "last_digit":   row[3],
            "venue":        row[4],
            "contribution": row[5],
            "problem":      row[6],
            "key_idea":     row[7],
            "achievement":  row[8],
            "autonomy":     row[9],
            "worth":        row[10],
            "reason":       row[11],
        })

    print(f"[INFO] 発表者リスト読み込み完了: {len(presenters)} 名")
    return presenters


# ─────────────────────────────────────────
#  ユーティリティ: 動画から音声 numpy 配列を抽出
# ─────────────────────────────────────────
def _extract_audio(video_path: Path) -> "np.ndarray":
    """
    PyAV で動画から 16kHz モノラル float32 の numpy 配列を返す。
    .mov / .mkv など多様な形式に対応するため、複数の方法でストリームを検出する。
    """
    try:
        import av
        import numpy as np
    except ImportError as e:
        print(f"[ERROR] 必要なライブラリが未インストールです: {e}")
        print("  pip install openai-whisper av numpy")
        sys.exit(1)

    container = av.open(str(video_path))

    # 方法1: container.streams.audio (PyAV の正式 API)
    audio_stream = None
    if container.streams.audio:
        audio_stream = container.streams.audio[0]

    # 方法2: 全ストリームをスキャン（フォールバック）
    if audio_stream is None:
        for s in container.streams:
            if hasattr(s, 'type') and s.type == "audio":
                audio_stream = s
                break

    if audio_stream is None:
        all_types = [getattr(s, 'type', 'unknown') for s in container.streams]
        raise ValueError(
            f"音声トラックが見つかりません: {video_path.name}\n"
            f"  検出されたストリーム: {all_types}\n"
            f"  -> 動画ファイルに音声が含まれていない可能性があります"
        )

    resampler = av.AudioResampler(format="flt", layout="mono", rate=16000)
    frames = []
    try:
        for frame in container.decode(audio_stream):
            for rf in resampler.resample(frame):
                frames.append(rf.to_ndarray().flatten())
    except av.AVError as e:
        if frames:
            print(f"    [WARN] デコード中にエラー（途中まで取得済み）: {e}")
        else:
            raise

    if not frames:
        raise ValueError(f"音声フレームをデコードできませんでした: {video_path.name}")

    return np.concatenate(frames)


def _is_garbage_transcript(text: str) -> bool:
    """
    無音区間の誤検出の全体判定。
    内容の 75% 以上が NOISE_WORDS の場合、または空文字の場合に True。
    """
    words = text.strip().lower().split()
    if not words:
        return True
    noise_count = sum(1 for w in words if w.strip('.,!?') in NOISE_WORDS)
    return noise_count / len(words) >= GARBAGE_RATIO


def _trim_trailing_noise(text: str) -> str:
    """
    末尾の無音誤認識（'you you you...'など）を除去する。
    最後の TRIM_WINDOW 単語を進む方向にスキャンし、
    NOISE_WORDS が連続する先頭位置に切り捨てる。
    """
    words = text.split()
    if not words:
        return text

    # 末尾から顧みてノイズでない最後の単語位置を探す
    cut = len(words)
    # 少なくとも TRIM_WINDOW 単語を移動ウィンドウでチェック
    window = min(TRIM_WINDOW, len(words))
    for i in range(len(words) - 1, len(words) - 1 - window, -1):
        if words[i].strip('.,!?').lower() not in NOISE_WORDS:
            cut = i + 1
            break
    else:
        # ウィンドウ全体がノイズの場合、更に遥かに山成りに探す
        cut = 0
        for i in range(len(words) - 1, -1, -1):
            if words[i].strip('.,!?').lower() not in NOISE_WORDS:
                cut = i + 1
                break

    trimmed = ' '.join(words[:cut]).rstrip()
    if cut < len(words):
        removed = len(words) - cut
        print(f"    [トリム] 末尾ノイズ {removed} 単語を除去しました")
    return trimmed


# ─────────────────────────────────────────
#  動画ファイルのグループ化（分割動画を統合）
# ─────────────────────────────────────────
def group_video_files(movie_dir: Path) -> list:
    """
    動画ファイルを「グループ」単位でまとめる。

    命名規則:
      <base>-<part>.<ext>  → 同じ <base> でグループ化して音声結合
        例: 8-1.mov, 8-2.mov, 8-3.mov → グループ名 "8"
      <base>.<ext>         → 単独グループ
        例: 9_maeno.mkv    → グループ名 "9_maeno"

    戻り値: [{"group": str, "files": [Path, ...]}, ...]
    """
    import re as _re
    all_files = []
    for ext in VIDEO_EXTENSIONS:
        all_files.extend(movie_dir.glob(f"*{ext}"))
    all_files = sorted(set(all_files))

    groups = {}  # group_name -> [Path]
    split_pattern = _re.compile(r'^(.+?)-(\d+)$')  # e.g. "8-1" -> base="8", part="1"

    for f in all_files:
        m = split_pattern.match(f.stem)
        if m:
            base = m.group(1)  # "8"
            if base not in groups:
                groups[base] = []
            groups[base].append(f)
        else:
            # 単独ファイル（グループ名 = stem）
            name = f.stem
            if name not in groups:
                groups[name] = []
            groups[name].append(f)

    result = []
    for gname, files in sorted(groups.items()):
        result.append({"group": gname, "files": sorted(files)})
    return result


# ─────────────────────────────────────────
#  動画（グループ） → 文字起こし (Whisper)
# ─────────────────────────────────────────
def transcribe_group(group_name: str, video_files: list, output_path: Path,
                     model_name: str = WHISPER_MODEL) -> str:
    """
    1本または複数本の動画を結合して文字起こし。
    出力ファイルが既に存在し、ゴミ判定されない場合はスキップ。
    """
    try:
        import numpy as np
        import whisper
    except ImportError as e:
        print(f"[ERROR] 必要なライブラリが未インストールです: {e}")
        print("  pip install openai-whisper av numpy")
        sys.exit(1)

    # 既存ファイルのチェック：ゴミでなければスキップ
    if output_path.exists():
        existing = output_path.read_text(encoding="utf-8")
        if not _is_garbage_transcript(existing):
            label = "+".join(f.name for f in video_files)
            print(f"  [SKIP] 文字起こし済み: {output_path.name}")
            return existing
        else:
            print(f"  [REDO] ゴミ検出（無音誤認識）のため再処理: {output_path.name}")
            output_path.unlink()

    label = " + ".join(f.name for f in video_files)
    print(f"  [Whisper] 文字起こし中: {label}")
    if len(video_files) > 1:
        print(f"    -> {len(video_files)} 本を結合して処理します")

    # 音声を順番に抽出・結合（音声なしファイルはスキップ）
    audio_chunks = []
    skipped = []
    for vf in video_files:
        print(f"    読み込み中: {vf.name}")
        try:
            audio_chunks.append(_extract_audio(vf))
        except ValueError as e:
            print(f"    [SKIP] 音声なしのためスキップ: {vf.name}")
            print(f"           ({str(e).splitlines()[0]})")
            skipped.append(vf.name)

    if not audio_chunks:
        raise ValueError(
            f"グループ内の全ファイルに音声がありません: {[f.name for f in video_files]}\n"
            f"  -> 録画設定を確認してください（マイクがオフになっていた可能性があります）"
        )
    if skipped:
        print(f"    [INFO] {len(skipped)} 本をスキップ、{len(audio_chunks)} 本で文字起こしを続行")

    audio_data = np.concatenate(audio_chunks)

    # Whisper モデルのロード
    local_pt = BASE_DIR / "Reference" / "base.pt"
    model_path = str(local_pt) if local_pt.exists() else model_name
    model = whisper.load_model(model_path)

    # 英語プレゼンを想定
    result = model.transcribe(audio_data, verbose=True, language="en")
    text = result["text"]

    # 末尾ノイズトリミング（セッション終了後の "you you you..." を除去）
    text = _trim_trailing_noise(text)

    # ゴミチェック（トリム後も全体がノイズなら警告）
    if _is_garbage_transcript(text):
        print(f"  [WARN] 文字起こし結果が無音誤認識の可能性があります: {output_path.name}")
        print(f"         内容: {text[:100]!r}")
        print(f"         -> 動画に音声が含まれているか確認してください")

    output_path.write_text(text, encoding="utf-8")
    print(f"  [DONE] 保存完了: {output_path.name} ({len(text)} 文字)")
    return text


# ─────────────────────────────────────────
#  評価プロンプト生成
# ─────────────────────────────────────────
def build_evaluation_prompt(transcript: str, presenter_name_list: list, slide_info: str = "") -> str:
    names_str = ", ".join(presenter_name_list) if presenter_name_list else "不明"
    excerpt = transcript[:6000] + ("..." if len(transcript) > 6000 else "")
    return f"""# Role
あなたは大学の学術プレゼンテーション（英語）の審査員です。

# Inputs
- transcript: {excerpt}
- slide_info: {slide_info or "(なし)"}
- presenter_name_list: {names_str}

# Task & Evaluation Criteria
以下の9項目を抽出・採点してください。
採点（0〜3点）は甘めに（原則2または3）。内容が破綻/全く理解できない場合のみ0か1。

1. Presenter's name: テキスト内で名乗っている名前。リストと照合。
2. Presenter's last digit: 冒頭や挨拶に出てくる発表者の末尾の数字。不明なら「不明」。
3. Paper Venue: 論文の会議名・ジャーナル名（例: CVPR, CHI, SIGGRAPH）。
4. Paper's contribution (0-3): 論文の貢献度（甘め採点）。
5. Problem setting (0-3): 課題設定の妥当性・面白さ（甘め採点）。
6. Key idea Achievement (0-3): 核心アイデアと達成度（甘め採点）。
7. Autonomy (0-3): 自分の言葉で語れているか（甘め採点）。
8. Worth listening? (0-3): 聞く価値があったか（甘め採点）。
9. Why: 日本語で一言二言（最大2文）の採点理由。

# Output Format
CSVのデータ行1行のみを出力してください（ヘッダー不要）。他の説明は出力しないでください。

[Name],[Digit],[Venue],[Score],[Score],[Score],[Score],[Score],[日本語理由]
"""


# ─────────────────────────────────────────
#  CSV 出力
# ─────────────────────────────────────────
HEADER = [
    "Presenter's name", "Presenter's last digit", "Paper Venue",
    "Paper's contribution", "Problem setting", "Key idea Achievement",
    "Autonomy", "Worth listening?", "Why do you evaluate so?"
]


def save_csv(rows: list, output_path: Path):
    write_header = not output_path.exists()
    with open(output_path, "a", encoding="utf-8-sig", newline="") as f:
        writer = csv.writer(f)
        if write_header:
            writer.writerow(HEADER)
        for row in rows:
            writer.writerow(row)
    print(f"[INFO] CSV保存: {output_path}")


# ─────────────────────────────────────────
#  .env から API キーを読み込む
# ─────────────────────────────────────────
def load_api_key() -> str:
    """環境変数または .env ファイルから GOOGLE_API_KEY を取得する。"""
    key = os.environ.get("GOOGLE_API_KEY")
    if key:
        return key
    env_file = BASE_DIR / ".env"
    if env_file.exists():
        for line in env_file.read_text(encoding="utf-8").splitlines():
            line = line.strip()
            if line.startswith("GOOGLE_API_KEY="):
                return line.split("=", 1)[1].strip()
    return ""


# ─────────────────────────────────────────
#  セッション全体向け評価プロンプト（複数発表者）
# ─────────────────────────────────────────
def build_session_prompt(transcript: str, presenters_in_session: list,
                          slide_info: str = "") -> str:
    """
    1 セッション（複数発表者）分の評価プロンプト。
    Gemini にトランスクリプト内の全発表者を自動特定させ
    各自の CSV 行を出力させる。
    """
    names_str = "\n".join(f"  - {p['name']}" for p in presenters_in_session)
    excerpt = transcript[:12000] + ("\n[...continues...]" if len(transcript) > 12000 else "")
    return f"""# Role
You are a strict academic presentation judge at a graduate university (English presentations).
Multiple student presentations were recorded continuously in one session.
Evaluate EACH presenter found in the transcript below.

# Pre-registered presenter list for this session:
{names_str}

# Transcript:
{excerpt}

# Slide info:
{slide_info or '(none)'}

# Evaluation Criteria
For each presenter found in the transcript, extract and score the following 9 items.
Scoring 0-3: Be GENEROUS - default is 2 or 3. Use 0 or 1 ONLY if content is completely incoherent.

1. Presenter name: Match to the pre-registered list. Use exact name from list.
2. Last digit: Last digit of student ID mentioned at start. Write "?" if not mentioned.
3. Paper Venue: Conference/journal (e.g., CVPR, CHI, SIGGRAPH, ICCV, NeurIPS, ECCV, AAAI).
4. Paper contribution (0-3): Contribution of the paper. Be generous.
5. Problem setting (0-3): Problem validity/interest. Be generous.
6. Key idea achievement (0-3): Core idea clarity. Be generous.
7. Autonomy (0-3): Explains in own words. Be generous.
8. Worth listening (0-3): Worth attending. Be generous.
9. Why (Japanese): 1-2 sentences maximum in Japanese explaining the score.

# Output Rules
- Output ONLY CSV data rows, one row per presenter FOUND in transcript.
- NO header. NO explanations. NO markdown fences. Plain CSV only.
- Skip any presenter from the list who does NOT appear in the transcript.
- Format exactly: Name,Digit,Venue,Contribution,Problem,KeyIdea,Autonomy,Worth,WhyJapanese
- Example: Tanaka Taro,5,CVPR,3,2,3,2,3,論文の貢献が明確で発表も分かりやすかった。

CSV rows:"""


# ─────────────────────────────────────────
#  Gemini API 呼び出し
# ─────────────────────────────────────────
def call_gemini_api(prompt: str, api_key: str,
                    model_name: str = "gemini-2.5-flash",
                    max_retries: int = 4) -> str:
    """Gemini API を呼び出してテキスト応答を返す。429 時は指数バックオフでリトライ。"""
    import time
    try:
        import google.generativeai as genai
    except ImportError:
        print("[ERROR] google-generativeai が未インストールです")
        print("  pip install google-generativeai")
        sys.exit(1)

    genai.configure(api_key=api_key)
    model = genai.GenerativeModel(model_name)

    for attempt in range(max_retries):
        try:
            response = model.generate_content(prompt)
            return response.text.strip()
        except Exception as e:
            err_str = str(e)
            if "429" in err_str or "quota" in err_str.lower():
                # 待機時間を抽出（retry_delay があれば使う、なければ指数バックオフ）
                import re as _re
                m = _re.search(r'retry_delay\s*\{\s*seconds:\s*(\d+)', err_str)
                wait = int(m.group(1)) + 2 if m else (10 * (2 ** attempt))
                wait = min(wait, 120)  # 最大120秒
                if attempt < max_retries - 1:
                    print(f"  [RATE LIMIT] レート制限。{wait}秒後にリトライ ({attempt+1}/{max_retries-1})...")
                    time.sleep(wait)
                else:
                    raise
            else:
                raise
    raise RuntimeError("max retries exceeded")


def parse_csv_rows(raw_text: str) -> list:
    """Gemini の返答から有効な CSV 行を抽出する。"""
    import io
    text = re.sub(r'```[a-z]*\n?', '', raw_text).strip()
    rows = []
    reader = csv.reader(io.StringIO(text))
    for row in reader:
        if len(row) >= 9 and row[0].strip():
            if row[0].strip().lower() in ("presenter's name", "name", "presenter"):
                continue
            rows.append([cell.strip() for cell in row[:9]])
    return rows


def get_week_from_group(group_name: str) -> int:
    """グループ名から週番号を抽出する。'8' -> 8, '9_maeno' -> 9"""
    m = re.match(r'^(\d+)', group_name)
    return int(m.group(1)) if m else 0


# ─────────────────────────────────────────
#  Gemini API 自動評価
# ─────────────────────────────────────────
def evaluate_with_api(api_key: str, inter_request_wait: int = 10):
    """
    全文字起こしファイルを Gemini API で評価し evaluation_results.csv に保存する。
    - 発表者1名ずつ個別にAPIコール（プロンプトを小さくしてQuota節約）
    - 1行取得するたびに即座にCSVへ追記（途中再開対応）
    - リクエスト間に inter_request_wait 秒のウェイト
    """
    import time
    from collections import defaultdict

    presenters = load_presenter_list(EVAL_SHEET)

    # 既存 CSV から評価済み名前を取得（途中再開用）
    evaluated_names = set()
    if OUTPUT_CSV.exists():
        with open(OUTPUT_CSV, encoding="utf-8-sig") as f:
            for row in csv.reader(f):
                if row and row[0] and row[0] != HEADER[0]:
                    evaluated_names.add(row[0].strip().lower())
    if evaluated_names:
        print(f"[INFO] 既存 CSV で {len(evaluated_names)} 名分の評価を確認済み（スキップ）")

    # CSV ヘッダーを初回のみ書き込む
    if not OUTPUT_CSV.exists():
        with open(OUTPUT_CSV, "w", encoding="utf-8-sig", newline="") as f:
            csv.writer(f).writerow(HEADER)

    # 週ごとに発表者を分類
    week_map = defaultdict(list)
    for pr in presenters:
        if pr["week"]:
            week_map[int(pr["week"])].append(pr)

    transcript_files = sorted(PRES_TRANS.glob("*_transcript.txt"))
    if not transcript_files:
        print(f"[WARN] 文字起こしファイルが見つかりません: {PRES_TRANS}")
        return

    prompts_dir = BASE_DIR / "Reference" / "evaluation_prompts"
    prompts_dir.mkdir(exist_ok=True)

    total_saved = 0
    for tf in transcript_files:
        text = tf.read_text(encoding="utf-8").strip()
        group_name = tf.stem.replace("_transcript", "")
        week_num = get_week_from_group(group_name)

        if not text or _is_garbage_transcript(text):
            print(f"\n  [SKIP] 有効なトランスクリプトなし: {tf.name}")
            continue

        presenters_in_week = week_map.get(week_num, [])
        if not presenters_in_week:
            print(f"\n  [WARN] 週 {week_num} の発表者リストなし: {tf.name}")
            continue

        unevaluated = [
            pr for pr in presenters_in_week
            if pr["name"].strip().lower() not in evaluated_names
        ]
        if not unevaluated:
            print(f"\n  [SKIP] 週 {week_num} は全員評価済み: {tf.name}")
            continue

        # 先頭ノイズをスキップして実際の発話開始位置を特定
        speech_start = 0
        for marker in ["hello", "good morning", "good afternoon", "hi everyone", "my name"]:
            pos = text.lower().find(marker)
            if pos != -1:
                speech_start = max(0, pos - 100)
                break
        speech_text = text[speech_start:]

        print(f"\n{'='*60}")
        print(f"[週 {week_num}] {tf.name} - 未評価 {len(unevaluated)} 名")
        print(f"{'='*60}")
        print(f"  発話開始: {speech_start} 文字目 / 全 {len(text)} 文字")


        def extract_presenter_section(full_text: str, presenter_name: str,
                                       all_presenters: list, window: int = 4500) -> str:
            """
            トランスクリプトから発表者の名前付近を抽出する。
            1. フルネーム → 姓のみ → 名のみ の順で検索
            2. 見つからない場合は全発表者数でセクションを均等分割し、
               順番に基づいた推定位置から抽出する
            """
            text_lower = full_text.lower()
            parts = presenter_name.split()
            # 検索候補: 姓優先（全角も考慮）
            candidates = [presenter_name.lower()] + [p.lower() for p in parts if len(p) > 2]

            best_pos = -1
            for cand in candidates:
                pos = text_lower.find(cand)
                if pos != -1:
                    best_pos = pos
                    break

            if best_pos != -1:
                start = max(0, best_pos - 200)
                end = min(len(full_text), best_pos + window)
                prefix = "[...] " if start > 0 else ""
                suffix = " [...]" if end < len(full_text) else ""
                return prefix + full_text[start:end] + suffix

            # 名前が見つからない場合: 順番で均等分割
            n = len(all_presenters)
            if n == 0:
                return full_text[:window]
            try:
                order = [p["name"] for p in all_presenters].index(presenter_name)
            except ValueError:
                order = 0
            section_len = len(full_text) // n
            start = max(0, order * section_len - 200)
            end = min(len(full_text), start + window)
            return f"[name not found - estimated section {order+1}/{n}]\n" + full_text[start:end]


        for idx, pr in enumerate(unevaluated):
            name = pr["name"]
            print(f"\n  [{idx+1}/{len(unevaluated)}] {name} を評価中...")

            # 発表者名でトランスクリプトから該当箇所を抽出（ノイズスキップ後のテキスト使用）
            excerpt = extract_presenter_section(speech_text, name, presenters_in_week)
            print(f"    抽出: {len(excerpt)} 文字")


            # 1名専用プロンプト（コンパクト）
            prompt = f"""You are a university academic presentation judge (English presentations).
Evaluate ONE specific presenter from the transcript of a multi-presenter session.

# Target Presenter: {name}
# All presenters in this session (for reference): {', '.join(p['name'] for p in presenters_in_week)}

# Transcript section (the part where {name} presents):
{excerpt}

# Scoring (0-3, be GENEROUS - default 2 or 3, use 0-1 only if completely incoherent):
1. Name: exact name matching "{name}" from the list
2. LastDigit: last digit of student ID mentioned. "?" if not found
3. Venue: paper conference/journal (CVPR, NeurIPS, ECCV, ICCV, SIGGRAPH, CHI, AAAI, etc.)
4. Contribution (0-3): paper contribution
5. Problem (0-3): problem setting quality
6. KeyIdea (0-3): key idea clarity
7. Autonomy (0-3): explains in own words
8. Worth (0-3): worth listening to
9. Why: 1-2 sentences in Japanese

Output EXACTLY ONE CSV line (no header, no explanation, no markdown):
Name,Digit,Venue,Contribution,Problem,KeyIdea,Autonomy,Worth,WhyJapanese"""


            try:
                raw = call_gemini_api(prompt, api_key)
                rows = parse_csv_rows(raw)

                if rows:
                    row = rows[0]
                    # CSV に即座に追記
                    with open(OUTPUT_CSV, "a", encoding="utf-8-sig", newline="") as f:
                        csv.writer(f).writerow(row)
                    evaluated_names.add(name.strip().lower())
                    total_saved += 1
                    print(f"    ✓ 保存: {row[0]} | {row[2]} | {row[3]},{row[4]},{row[5]},{row[6]},{row[7]}")
                    print(f"      理由: {row[8][:50] if len(row) > 8 else '?'}")
                else:
                    # 生の応答を保存して手動確認
                    raw_file = prompts_dir / f"{group_name}_{name.replace(' ','_')}_response.txt"
                    raw_file.write_text(raw, encoding="utf-8")
                    print(f"    [WARN] CSVパース失敗。応答保存: {raw_file.name}")
                    print(f"    応答: {raw[:120]}")

            except Exception as e:
                print(f"    [ERROR] {name}: {e}")
                print("    -> 次の発表者へ")

            # リクエスト間のウェイト（最後の1件は不要）
            if idx < len(unevaluated) - 1:
                print(f"    ({inter_request_wait}秒待機...)")
                time.sleep(inter_request_wait)

        print(f"\n  [週 {week_num}] 完了 - ここまで累計 {total_saved} 名を保存")

    print(f"\n{'='*60}")
    print(f"[INFO] 全評価完了。合計 {total_saved} 名を {OUTPUT_CSV} に保存しました")
    print(f"{'='*60}")




# ─────────────────────────────────────────
#  コマンド: リスト表示
# ─────────────────────────────────────────
def list_presenters():
    presenters = load_presenter_list(EVAL_SHEET)
    print(f"\n{'週':>4} {'番号':>4}  {'名前':<40} {'評価'}")
    print("-" * 60)
    for p in presenters:
        filled = "[OK]" if (p["venue"] or p["contribution"] is not None) else " "
        print(f"  {str(p['week']):>4} {str(p['number']):>4}  {p['name']:<40} {filled}")
    unevaluated = [p for p in presenters if p["venue"] is None and p["contribution"] is None]
    print(f"\n計: {len(presenters)} 名  |  未評価: {len(unevaluated)} 名")


# ─────────────────────────────────────────
#  コマンド: 文字起こし（分割動画の結合対応）
# ─────────────────────────────────────────
def transcribe_all(model_name: str = WHISPER_MODEL):
    presenters = load_presenter_list(EVAL_SHEET)
    name_list = [p["name"] for p in presenters]

    groups = group_video_files(PRES_MOVIES)
    if not groups:
        print(f"\n[WARN] 動画ファイルが見つかりません: {PRES_MOVIES}")
        print(f"  対応形式: {', '.join(VIDEO_EXTENSIONS)}")
        print("  命名規則:")
        print(f"    {PRES_MOVIES}/<発表者名>.<ext>           # 単独")
        print(f"    {PRES_MOVIES}/<週>-<番号>.<ext>          # 分割 (例: 8-1.mov, 8-2.mov)")
        print(f"\n  Evaluation Sheet.xlsx に登録済みの発表者（{len(name_list)}名）:")
        for name in name_list[:10]:
            print(f"    - {name}")
        if len(name_list) > 10:
            print(f"    ... 他 {len(name_list)-10} 名")
        return []

    # グループ一覧を表示
    print(f"\n[INFO] 動画グループ {len(groups)} 件を処理します")
    for g in groups:
        parts = [f.name for f in g["files"]]
        if len(parts) == 1:
            print(f"  [{g['group']}] {parts[0]}")
        else:
            print(f"  [{g['group']}] 分割 {len(parts)} 本: {' + '.join(parts)}")

    print()
    results = []
    for g in groups:
        group_name = g["group"]
        files = g["files"]
        transcript_path = PRES_TRANS / f"{group_name}_transcript.txt"
        try:
            text = transcribe_group(group_name, files, transcript_path, model_name)
            results.append({"group": group_name, "files": files,
                            "transcript": text, "path": transcript_path})
        except Exception as e:
            print(f"  [ERROR] グループ [{group_name}]: {e}")
    return results


# ─────────────────────────────────────────
#  コマンド: 評価プロンプト生成
# ─────────────────────────────────────────
def evaluate_all():
    presenters = load_presenter_list(EVAL_SHEET)
    name_list = [p["name"] for p in presenters]

    transcript_files = sorted(PRES_TRANS.glob("*_transcript.txt"))
    if not transcript_files:
        print(f"\n[WARN] 文字起こしファイルが見つかりません: {PRES_TRANS}")
        print("  先に文字起こしを実施: python evaluate_presentations.py --transcribe-only")
        return

    prompts_dir = BASE_DIR / "Reference" / "evaluation_prompts"
    prompts_dir.mkdir(exist_ok=True)

    print(f"\n[INFO] {len(transcript_files)} 件の文字起こしを処理します")
    for tf in transcript_files:
        text = tf.read_text(encoding="utf-8")
        stem = tf.stem.replace("_transcript", "")

        # 発表者名の推定（ファイル名との部分一致）
        matched = None
        for name in name_list:
            normalized = name.replace(" ", "_").upper()
            if normalized in stem.upper() or stem.upper() in normalized:
                matched = name
                break
        presenter_hint = [matched] if matched else name_list

        prompt = build_evaluation_prompt(text, presenter_hint)
        prompt_file = prompts_dir / (stem + "_prompt.txt")
        prompt_file.write_text(prompt, encoding="utf-8")
        print(f"  [PROMPT] {prompt_file.name}")

    print(f"\n[INFO] 評価プロンプトを {prompts_dir} に保存しました")
    print("  各プロンプトを LLM（ChatGPT / Gemini 等）に貼り付けて評価結果を取得し、")
    print(f"  {OUTPUT_CSV} に行を追記してください。")
    print()
    print("  ※ GOOGLE_API_KEY 等を設定して --api フラグを追加すれば自動評価も可能です。")


# ─────────────────────────────────────────
#  エントリポイント
# ─────────────────────────────────────────
def main():
    parser = argparse.ArgumentParser(
        description="プレゼン評価パイプライン",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
使い方:
  python evaluate_presentations.py --list              発表者リスト表示
  python evaluate_presentations.py --transcribe-only   動画 -> 文字起こし
  python evaluate_presentations.py --evaluate-only     文字起こし -> 評価プロンプト生成
  python evaluate_presentations.py                     文字起こし + 評価プロンプト生成（一括）
  python evaluate_presentations.py --model medium      Whisper モデルを変更

動画の置き場所 (Reference/presentation_movies/):
  単独: <名前>.mp4 / .mov / .mkv / .avi など
  分割: <週>-1.mov, <週>-2.mov, ... → 自動結合して <週>_transcript.txt を生成
  例: 8-1.mov, 8-2.mov, 8-3.mov → 8_transcript.txt
"""
    )
    parser.add_argument("--list",            action="store_true", help="発表者リストを表示")
    parser.add_argument("--transcribe-only", action="store_true", help="文字起こしのみ")
    parser.add_argument("--evaluate-only",   action="store_true", help="評価プロンプト生成のみ")
    parser.add_argument("--api",             action="store_true",
                        help="Gemini API で自動評価して evaluation_results.csv に保存")
    parser.add_argument("--model",           default=WHISPER_MODEL,
                        help=f"Whisper モデル (default: {WHISPER_MODEL})")
    args = parser.parse_args()

    if args.list:
        list_presenters()
    elif args.api:
        api_key = load_api_key()
        if not api_key:
            print("[ERROR] GOOGLE_API_KEY が設定されていません")
            print("  .env ファイルに 'GOOGLE_API_KEY=...' を記載するか")
            print("  環境変数 GOOGLE_API_KEY を設定してください")
            sys.exit(1)
        print(f"[INFO] API キーを読み込みました: {'*' * 20}{api_key[-4:]}")
        evaluate_with_api(api_key)
    elif args.transcribe_only:
        transcribe_all(args.model)
    elif args.evaluate_only:
        evaluate_all()
    else:
        results = transcribe_all(args.model)
        if results:
            evaluate_all()
        else:
            print("\n[INFO] 動画が用意されたら再実行してください。")


if __name__ == "__main__":
    main()
